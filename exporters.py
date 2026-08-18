"""ReportData -> XLSX / HTML / 메일 본문 변환."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from report_engine import fmt_won, fmt_int, fmt_pct

TITLE_FILL = PatternFill("solid", fgColor="1F2937")
SECTION_FILL = PatternFill("solid", fgColor="EEF2FF")
HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
TOTAL_FILL = PatternFill("solid", fgColor="FEF3C7")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _title(ws, cell, text, size=14, bold=True, color="FFFFFF", fill=None):
    ws[cell] = text
    ws[cell].font = Font(size=size, bold=bold, color=color)
    if fill:
        ws[cell].fill = fill


def _section(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(size=11, bold=True, color="4338CA")


def _header_row(ws, row, start_col, labels):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=label)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")


def _write_row(ws, row, start_col, values, number_formats=None):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.border = BORDER
        if number_formats and number_formats[i]:
            c.number_format = number_formats[i]


def _autosize(ws, min_width=8, max_width=42):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            widths[col] = max(widths.get(col, min_width), min(max_width, len(str(cell.value)) + 2))
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_workbook(rd, meta) -> bytes:
    wb = Workbook()
    _sheet_summary(wb.active, rd, meta)
    wb.active.title = "① 종합 요약"
    _sheet_room(wb.create_sheet("② 시설별 객실 분석"), rd)
    _sheet_matrix(wb.create_sheet("③ 패키지×객실 매트릭스"), rd)
    _sheet_adr(wb.create_sheet("④ ADR 분석"), rd)
    _sheet_detail(wb.create_sheet("⑤ 체크인 일자별"), rd)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_summary(ws, rd, meta):
    _title(ws, "B1", f"{meta['influencer']} × {meta['hotel']} — 공동구매 매출 실적 요약", size=14)
    ws["B2"] = f"판매기간: {meta['sale_period']}   |   투숙기간: {meta['stay_period']}"
    ws["B2"].font = Font(size=10, color="6B7280")

    _header_row(ws, 4, 2, ["확정 건수(확정+완료)", "확정 박수", "확정 매출액(원)", "취소 건수", "취소율"])
    _write_row(ws, 5, 2, [
        rd.confirmed_count, rd.confirmed_nights, rd.confirmed_revenue,
        rd.cancelled_count, rd.cancel_rate,
    ], number_formats=[None, None, "#,##0", None, "0.0%"])

    row = 7
    _section(ws, f"B{row}", "▶  전체 예약 현황")
    row += 1
    _header_row(ws, row, 2, ["숙소", "상태", "건 수", "박 수", "매출액 (원)"])
    row += 1
    for _, r in rd.overall_status.iterrows():
        _write_row(ws, row, 2, [
            r["숙소"], r["상태"], r["건수"], r["박수"], r["매출액"],
        ], number_formats=[None, None, None, None, "#,##0"])
        if r["상태"] == "전체":
            for col in range(2, 7):
                ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1

    row += 1
    _section(ws, f"B{row}", "▶  패키지별 분석 (확정 기준)")
    row += 1
    _header_row(ws, row, 2, ["패키지", "건 수", "박 수", "매출액 (원)", "ADR(박)", "ADR(건)", "취소건", "취소율", "매출비중"])
    row += 1
    for _, r in rd.package_analysis.iterrows():
        fill = TOTAL_FILL if r["패키지"] == "합계" else None
        _write_row(ws, row, 2, [
            r["패키지"], r["건수"], r["박수"], r["매출액"], r["ADR_박"], r["ADR_건"],
            r["취소건"], r["취소율"], r["매출비중"],
        ], number_formats=[None, None, None, "#,##0", "#,##0", "#,##0", None, "0.0%", "0.0%"])
        if fill:
            for col in range(2, 11):
                ws.cell(row=row, column=col).fill = fill
        row += 1

    row += 1
    _section(ws, f"B{row}", "▶  체크인 월별 현황 (확정 기준)")
    row += 1
    header_row_idx = row
    ws.cell(row=row, column=2, value="구분")
    col = 3
    for m in rd.months:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
        c = ws.cell(row=row, column=col, value=f"{m.year}년 {m.month:02d}월")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        col += 3
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    ws.cell(row=row, column=col, value="합계").font = Font(bold=True)
    row += 1
    labels = ["건", "박", "매출 (원)"] * (len(rd.months) + 1)
    _header_row(ws, row, 3, labels)
    row += 1
    for label in ["전체", "확정", "취소"]:
        vals = ["전체" if label == "전체" else label]
        totals = [0, 0, 0]
        col = 3
        line = [vals[0]]
        for m in rd.months:
            cnt, nights, rev = rd.monthly_checkin[label][m]
            line += [cnt, nights, rev]
            totals[0] += cnt; totals[1] += nights; totals[2] += rev
        line += totals
        _write_row(ws, row, 2, line, number_formats=[None] + ([None, None, "#,##0"] * (len(rd.months) + 1)))
        row += 1

    row += 1
    _section(ws, f"B{row}", f"▶  일별 판매 현황 (판매기간: {meta['sale_period']})")
    row += 1
    _header_row(ws, row, 2, ["날짜", "요일", "건 수", "박 수", "매출액 (원)", "누적 매출액 (원)"])
    row += 1
    for _, r in rd.daily_sales.iterrows():
        _write_row(ws, row, 2, [
            r["판매일"].strftime("%m-%d"), r["요일"], r["건수"], r["박수"], r["매출액"], r["누적매출액"],
        ], number_formats=[None, None, None, None, "#,##0", "#,##0"])
        row += 1
    _write_row(ws, row, 2, ["합계", "", rd.total_count, rd.total_nights, rd.total_revenue, ""],
                number_formats=[None, None, None, None, "#,##0", None])
    row += 2

    _section(ws, f"B{row}", "▶  체크인 요일별 현황 (확정 기준)")
    row += 1
    _header_row(ws, row, 2, ["요일", "건 수", "박 수", "매출액 (원)", "ADR (원/박)", "비중 (매출)"])
    row += 1
    for _, r in rd.weekday_checkin.iterrows():
        _write_row(ws, row, 2, [r["요일"], r["건수"], r["박수"], r["매출액"], r["ADR"], r["비중"]],
                    number_formats=[None, None, None, "#,##0", "#,##0", "0.0%"])
        row += 1
    _write_row(ws, row, 2, ["합계", rd.confirmed_count, rd.confirmed_nights, rd.confirmed_revenue,
                              rd.overall_adr, 1.0],
                number_formats=[None, None, None, "#,##0", "#,##0", "0.0%"])

    _autosize(ws)


def _sheet_room(ws, rd):
    _section(ws, "B1", "▸  전체")
    ws.merge_cells("C2:E2"); ws["C2"] = "전체 예약"
    ws.merge_cells("F2:H2"); ws["F2"] = "확정 예약"
    ws.merge_cells("I2:K2"); ws["I2"] = "취소 예약"
    for cell in ["C2", "F2", "I2"]:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal="center")
    _header_row(ws, 3, 2, ["객실명", "건", "박", "매출액 (원)", "건", "박", "매출액 (원)", "건", "박", "매출액 (원)"])
    row = 4
    for _, r in rd.room_analysis.iterrows():
        fill = TOTAL_FILL if r["객실명"] == "전체 소계" else None
        _write_row(ws, row, 2, [
            r["객실명"], r["전체_건수"], r["전체_박수"], r["전체_매출"],
            r["확정_건수"], r["확정_박수"], r["확정_매출"],
            r["취소_건수"], r["취소_박수"], r["취소_매출"],
        ], number_formats=[None, None, None, "#,##0", None, None, "#,##0", None, None, "#,##0"])
        if fill:
            for col in range(2, 11):
                ws.cell(row=row, column=col).fill = fill
        row += 1
    _autosize(ws)


def _sheet_matrix(ws, rd):
    ws["B1"] = "객실명"
    ws["B1"].font = Font(bold=True)
    col = 3
    for pkg in rd.matrix_packages + ["소계"]:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        c = ws.cell(row=1, column=col, value=pkg)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        col += 3
    labels = ["건", "박", "매출"] * (len(rd.matrix_packages) + 1)
    _header_row(ws, 2, 3, labels)
    row = 3
    for room in rd.matrix_rooms:
        fill = TOTAL_FILL if room == "합계" else None
        ws.cell(row=row, column=2, value=room).border = BORDER
        if fill:
            ws.cell(row=row, column=2).fill = fill
        col = 3
        for pkg in rd.matrix_packages + ["소계"]:
            cnt, nights, rev = rd.matrix[room].get(pkg, (0, 0, 0))
            for j, (val, fmt) in enumerate(zip([cnt, nights, rev], [None, None, "#,##0"])):
                cell = ws.cell(row=row, column=col + j, value=(val if val else None))
                cell.border = BORDER
                if fmt:
                    cell.number_format = fmt
                if fill:
                    cell.fill = fill
            col += 3
        row += 1
    _autosize(ws)


def _sheet_adr(ws, rd):
    ws["B1"] = "ADR (Average Daily Rate) 분석"
    ws["B1"].font = Font(size=13, bold=True)
    ws["B3"] = "전체 ADR"; ws["B3"].font = Font(bold=True)
    ws["B4"] = rd.overall_adr
    ws["B4"].number_format = "#,##0"

    row = 6
    _section(ws, f"B{row}", "▶  패키지별 ADR (확정 기준)")
    row += 1
    _header_row(ws, row, 2, ["패키지", "건 수", "박 수", "확정 매출", "ADR (원/박)", "건당 평균단가", "매출 비중", "취소율"])
    row += 1
    for _, r in rd.adr_package.iterrows():
        fill = TOTAL_FILL if r["패키지"] == "합계" else None
        _write_row(ws, row, 2, [
            r["패키지"], r["건수"], r["박수"], r["매출액"], r["ADR_박"], r["ADR_건"], r["매출비중"], r["취소율"],
        ], number_formats=[None, None, None, "#,##0", "#,##0", "#,##0", "0.0%", "0.0%"])
        if fill:
            for col in range(2, 10):
                ws.cell(row=row, column=col).fill = fill
        row += 1

    row += 1
    _section(ws, f"B{row}", "▶  객실별 ADR (확정 기준 / ADR 높은 순)")
    row += 1
    _header_row(ws, row, 2, ["객실명", "건 수", "박 수", "확정 매출", "ADR (원/박)", "건당 평균단가", "비중"])
    row += 1
    for _, r in rd.adr_room.iterrows():
        _write_row(ws, row, 2, [
            r["객실명"], r["건수"], r["박수"], r["확정매출"], r["ADR"], r["건당평균단가"], r["비중"],
        ], number_formats=[None, None, None, "#,##0", "#,##0", "#,##0", "0.0%"])
        row += 1
    _autosize(ws)


def _sheet_detail(ws, rd):
    _section(ws, "B1", "▶  체크인 일자별 현황 (확정 기준)")
    _header_row(ws, 2, 2, ["체크인", "요일", "객실명", "패키지", "건 수", "박 수", "매출액 (원)"])
    row = 3
    for _, r in rd.checkin_detail.iterrows():
        _write_row(ws, row, 2, [
            r["체크인"], r["요일"], r["객실명"], r["패키지"], r["건수"], r["박수"], r["매출액"],
        ], number_formats=[None, None, None, None, None, None, "#,##0"])
        row += 1
    _write_row(ws, row, 2, ["합계", None, None, None, rd.confirmed_count, rd.confirmed_nights, rd.confirmed_revenue],
                number_formats=[None, None, None, None, None, None, "#,##0"])
    _autosize(ws)


# ---------------- HTML ----------------

_HTML_STYLE = """
<style>
  body { font-family: -apple-system, "Malgun Gothic", sans-serif; color:#111827; background:#F9FAFB; margin:0; padding:32px; }
  .wrap { max-width: 1000px; margin: 0 auto; }
  h1 { font-size: 20px; margin-bottom:4px; }
  .subtitle { color:#6B7280; font-size:13px; margin-bottom:24px; }
  .kpi-row { display:flex; gap:12px; margin-bottom:28px; flex-wrap:wrap; }
  .kpi { background:#fff; border:1px solid #E5E7EB; border-radius:10px; padding:14px 18px; flex:1; min-width:140px; }
  .kpi .label { font-size:12px; color:#6B7280; }
  .kpi .value { font-size:20px; font-weight:700; margin-top:4px; }
  section { margin-bottom:32px; }
  h2 { font-size:15px; color:#4338CA; border-left:4px solid #4338CA; padding-left:8px; margin-bottom:12px; }
  table { border-collapse: collapse; width:100%; background:#fff; font-size:13px; }
  th, td { border:1px solid #E5E7EB; padding:6px 10px; text-align:right; }
  th { background:#F3F4F6; text-align:center; }
  td:first-child, th:first-child { text-align:left; }
  tr.total td { background:#FEF3C7; font-weight:700; }
  .table-scroll { overflow-x:auto; }
</style>
"""


def _df_to_html_rows(df, cols, headers, total_label_col=None, total_values=None, formats=None):
    out = ["<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"]
    for _, r in df.iterrows():
        cells = []
        for c, fmt in zip(cols, formats or [None] * len(cols)):
            v = r[c]
            cells.append(f"<td>{_fmt(v, fmt)}</td>")
        cls = " class='total'" if total_label_col and r.get(total_label_col) in ("합계", "전체 소계") else ""
        out.append(f"<tr{cls}>" + "".join(cells) + "</tr>")
    return "".join(out)


def render_overall_status_table(df) -> str:
    """rd.overall_status(숙소/상태/건수/박수/매출액)를 HTML 표로 변환. '전체' 상태 행은 볼드."""
    th = "border:1px solid #E5E7EB;padding:6px 10px;text-align:center;background:#F3F4F6;font-size:13px;"
    td_l = "border:1px solid #E5E7EB;padding:6px 10px;text-align:left;font-size:13px;"
    td_r = "border:1px solid #E5E7EB;padding:6px 10px;text-align:right;font-size:13px;"
    header = "<tr>" + "".join(f"<th style='{th}'>{h}</th>" for h in ["숙소", "상태", "건 수", "박 수", "매출액"]) + "</tr>"
    body = []
    for _, r in df.iterrows():
        bold = "font-weight:700;" if r["상태"] == "전체" else ""
        body.append(
            "<tr>"
            f"<td style='{td_l}{bold}'>{r['숙소']}</td>"
            f"<td style='{td_l}{bold}'>{r['상태']}</td>"
            f"<td style='{td_r}{bold}'>{fmt_int(r['건수'])}</td>"
            f"<td style='{td_r}{bold}'>{fmt_int(r['박수'])}</td>"
            f"<td style='{td_r}{bold}'>{fmt_won(r['매출액'])}</td>"
            "</tr>"
        )
    return "<table style='width:100%;border-collapse:collapse;'>" + header + "".join(body) + "</table>"


def _fmt(v, kind):
    if kind == "won":
        return fmt_won(v)
    if kind == "pct":
        return fmt_pct(v)
    if kind == "int":
        return fmt_int(v)
    return v


def build_html(rd, meta) -> str:
    overall_rows = render_overall_status_table(rd.overall_status)

    pkg_rows = _df_to_html_rows(
        rd.package_analysis, ["패키지", "건수", "박수", "매출액", "ADR_박", "ADR_건", "취소건", "취소율", "매출비중"],
        ["패키지", "건 수", "박 수", "매출액", "ADR(박)", "ADR(건)", "취소건", "취소율", "매출비중"],
        total_label_col="패키지",
        formats=[None, "int", "int", "won", "won", "won", "int", "pct", "pct"],
    )
    room_rows = _df_to_html_rows(
        rd.room_analysis, ["객실명", "전체_건수", "전체_박수", "전체_매출", "확정_건수", "확정_박수", "확정_매출", "취소_건수", "취소_박수", "취소_매출"],
        ["객실명", "전체 건", "전체 박", "전체 매출", "확정 건", "확정 박", "확정 매출", "취소 건", "취소 박", "취소 매출"],
        total_label_col="객실명",
        formats=[None, "int", "int", "won", "int", "int", "won", "int", "int", "won"],
    )
    adr_room_rows = _df_to_html_rows(
        rd.adr_room, ["객실명", "건수", "박수", "확정매출", "ADR", "건당평균단가", "비중"],
        ["객실명", "건 수", "박 수", "확정 매출", "ADR(원/박)", "건당 평균단가", "비중"],
        formats=[None, "int", "int", "won", "won", "won", "pct"],
    )
    daily_rows = "".join(
        f"<tr><td>{r['판매일'].strftime('%m-%d')}</td><td>{r['요일']}</td><td>{fmt_int(r['건수'])}</td>"
        f"<td>{fmt_int(r['박수'])}</td><td>{fmt_won(r['매출액'])}</td><td>{fmt_won(r['누적매출액'])}</td></tr>"
        for _, r in rd.daily_sales.iterrows()
    )
    weekday_rows = "".join(
        f"<tr><td>{r['요일']}</td><td>{fmt_int(r['건수'])}</td><td>{fmt_int(r['박수'])}</td>"
        f"<td>{fmt_won(r['매출액'])}</td><td>{fmt_won(r['ADR'])}</td><td>{fmt_pct(r['비중'])}</td></tr>"
        for _, r in rd.weekday_checkin.iterrows()
    )

    return f"""<!doctype html><html><head><meta charset="utf-8">
<title>{meta['influencer']} x {meta['hotel']} 실적 요약</title>
{_HTML_STYLE}
</head><body><div class="wrap">
<h1>{meta['influencer']} × {meta['hotel']} — 공동구매 매출 실적 요약</h1>
<div class="subtitle">판매기간: {meta['sale_period']} &nbsp;|&nbsp; 투숙기간: {meta['stay_period']}</div>

<div class="kpi-row">
  <div class="kpi"><div class="label">확정 건수(확정+완료)</div><div class="value">{fmt_int(rd.confirmed_count)}</div></div>
  <div class="kpi"><div class="label">확정 박수</div><div class="value">{fmt_int(rd.confirmed_nights)}</div></div>
  <div class="kpi"><div class="label">확정 매출액</div><div class="value">{fmt_won(rd.confirmed_revenue)}</div></div>
  <div class="kpi"><div class="label">취소 건수</div><div class="value">{fmt_int(rd.cancelled_count)}</div></div>
  <div class="kpi"><div class="label">취소율</div><div class="value">{fmt_pct(rd.cancel_rate)}</div></div>
</div>

<section><h2>전체 예약 현황</h2><div class="table-scroll">{overall_rows}</div></section>
<section><h2>패키지별 분석 (확정 기준)</h2><div class="table-scroll"><table>{pkg_rows}</table></div></section>
<section><h2>시설별 객실 분석</h2><div class="table-scroll"><table>{room_rows}</table></div></section>
<section><h2>일별 판매 현황</h2><div class="table-scroll"><table>
<tr><th>날짜</th><th>요일</th><th>건 수</th><th>박 수</th><th>매출액</th><th>누적 매출액</th></tr>
{daily_rows}
</table></div></section>
<section><h2>체크인 요일별 현황 (확정 기준)</h2><div class="table-scroll"><table>
<tr><th>요일</th><th>건 수</th><th>박 수</th><th>매출액</th><th>ADR</th><th>비중</th></tr>
{weekday_rows}
</table></div></section>
<section><h2>객실별 ADR (확정 기준 / 높은 순)</h2><div class="table-scroll"><table>{adr_room_rows}</table></div></section>
</div></body></html>"""


def build_email_body(rd, meta, sender_name="[담당자명]") -> str:
    return f"""안녕하세요.
트립비토즈 {sender_name}입니다.

{meta['sale_period_kr']}
진행된 {meta['influencer']} × {meta['hotel']} 공동구매 판매 실적을 아래와 같이 공유드립니다.

[실적 요약]
▪ 판매 기간: {meta['sale_period']}
▪ 투숙 기간: {meta['stay_period']}
▪ 판매 실적
- 전체 예약: {fmt_int(rd.total_count)}건 / {fmt_int(rd.total_nights)}박 / {fmt_won(rd.total_revenue)}
- 확정 예약: {fmt_int(rd.confirmed_count)}건 / {fmt_int(rd.confirmed_nights)}박 / {fmt_won(rd.confirmed_revenue)}
- 취소 건수: {fmt_int(rd.cancelled_count)}건 (취소율 {fmt_pct(rd.cancel_rate)})

세부 내역은 아래 첨부된 표를 참고 부탁드리며,
추가 문의 사항이 있으시면 언제든지 연락 주시기 바랍니다.

이번 공구에 적극적으로 협조해 주신 덕분에 좋은 성과를 거둘 수 있었습니다.
앞으로도 지속적인 협력 부탁드립니다.

감사합니다.

{sender_name} 드림
트립비토즈
"""
